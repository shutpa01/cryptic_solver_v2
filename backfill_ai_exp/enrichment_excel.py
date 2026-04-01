"""Export pending enrichments to Excel for review, then import decisions back.

Usage:
    python scripts/enrichment_excel.py export    # Create Excel file
    python scripts/enrichment_excel.py import    # Read decisions and apply
"""

import os
import sqlite3
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

CLUES_DB = os.path.join(ROOT, "data", "clues_master.db")
CRYPTIC_DB = os.path.join(ROOT, "data", "cryptic_new.db")
EXCEL_PATH = os.path.join(ROOT, "data", "enrichment_review.xlsx")


def export_to_excel():
    """Export pending_enrichments to Excel with a Decision column."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    conn = sqlite3.connect(CLUES_DB)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT id, type, word, letters
        FROM pending_enrichments
        ORDER BY type, word
    """).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enrichment Review"

    # Headers
    headers = ["ID", "Type", "Word", "Letters", "Decision"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r["id"])
        ws.cell(row=i, column=2, value=r["type"])
        ws.cell(row=i, column=3, value=r["word"]).font = Font(size=12)
        ws.cell(row=i, column=4, value=r["letters"]).font = Font(size=12, bold=True)
        # Decision column — leave blank, user types Y or N
        ws.cell(row=i, column=5, value="").font = Font(size=12)

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 12

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(EXCEL_PATH)
    print(f"Exported {len(rows)} pairs to {EXCEL_PATH}")
    print("Mark column E with Y (accept) or N (reject), then run: python scripts/enrichment_excel.py import")


def import_from_excel():
    """Read decisions from Excel and apply to DB."""
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    decisions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_id, etype, word, letters, decision = row
        if decision and str(decision).strip().upper() in ("Y", "N"):
            decisions.append((int(row_id), etype, word, letters, str(decision).strip().upper()))

    if not decisions:
        print("No decisions found. Mark column E with Y or N.")
        return

    accepted = 0
    rejected = 0

    clues_conn = sqlite3.connect(CLUES_DB, timeout=30)
    ref_conn = sqlite3.connect(CRYPTIC_DB, timeout=30)

    for row_id, etype, word, letters, decision in decisions:
        if decision == "Y":
            # Add to reference DB
            if etype == "synonym":
                existing = ref_conn.execute(
                    "SELECT 1 FROM synonyms_pairs WHERE LOWER(word)=? AND UPPER(synonym)=?",
                    (word.lower(), letters.upper())
                ).fetchone()
                if not existing:
                    ref_conn.execute(
                        "INSERT INTO synonyms_pairs (word, synonym, source) VALUES (?, ?, ?)",
                        (word.lower(), letters.upper(), "enrichment_batch")
                    )
            elif etype == "abbreviation":
                existing = ref_conn.execute(
                    "SELECT 1 FROM wordplay WHERE LOWER(indicator)=? AND UPPER(substitution)=?",
                    (word.lower(), letters.upper())
                ).fetchone()
                if not existing:
                    ref_conn.execute(
                        "INSERT OR IGNORE INTO wordplay (indicator, substitution) VALUES (?, ?)",
                        (word.lower(), letters.upper())
                    )
            elif etype == "definition":
                existing = ref_conn.execute(
                    "SELECT 1 FROM definition_answers_augmented WHERE LOWER(definition)=? AND UPPER(answer)=?",
                    (word.lower(), letters.upper())
                ).fetchone()
                if not existing:
                    ref_conn.execute(
                        "INSERT INTO definition_answers_augmented (definition, answer, source) VALUES (?, ?, ?)",
                        (word.lower(), letters.upper(), "enrichment_batch")
                    )
            # Remove from pending
            clues_conn.execute("DELETE FROM pending_enrichments WHERE id = ?", (row_id,))
            accepted += 1

        elif decision == "N":
            # Record rejection and remove from pending
            existing = clues_conn.execute(
                "SELECT 1 FROM rejected_enrichments WHERE type=? AND word=? AND letters=?",
                (etype, word, letters)
            ).fetchone()
            if not existing:
                clues_conn.execute(
                    "INSERT INTO rejected_enrichments (type, word, letters) VALUES (?, ?, ?)",
                    (etype, word, letters)
                )
            clues_conn.execute("DELETE FROM pending_enrichments WHERE id = ?", (row_id,))
            rejected += 1

    ref_conn.commit()
    ref_conn.close()
    clues_conn.commit()
    clues_conn.close()

    print(f"Done: {accepted} accepted, {rejected} rejected, {len(decisions)} total processed")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python scripts/enrichment_excel.py [export|import]")
        sys.exit(1)

    cmd = sys.argv[1].lower()
    if cmd == "export":
        export_to_excel()
    elif cmd == "import":
        import_from_excel()
    else:
        print(f"Unknown command: {cmd}")
