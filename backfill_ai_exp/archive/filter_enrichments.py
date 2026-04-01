"""Filter enrichment proposals mechanically, then export a cleaned Excel for review.

Auto-rejects:
1. Already in reference DB (synonyms_pairs or definition_answers_augmented)
2. Previously rejected
3. Single-letter mappings that aren't well-known abbreviations
4. Word == letters (circular)
5. Letters contained in word or word contained in letters (likely extraction artefact)
6. Very short words mapping to long values (suspicious)
7. Common words mapping to unrelated short values

Outputs: data/enrichment_review_filtered.xlsx
"""

import json
import os
import re
import sqlite3
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CRYPTIC_DB = os.path.join(ROOT, "data", "cryptic_new.db")
CLUES_DB = os.path.join(ROOT, "data", "clues_master.db")
INPUT_PATH = os.path.join(ROOT, "data", "enrichment_validated.jsonl")
OUTPUT_PATH = os.path.join(ROOT, "data", "enrichment_review_filtered.xlsx")

# Well-known single-letter cryptic abbreviations
KNOWN_ABBREVS = {
    "a": {"A"}, "about": {"C", "RE"}, "ace": {"A"}, "adult": {"A", "X"},
    "afternoon": {"PM"}, "against": {"V"}, "america": {"US"},
    "american": {"US"}, "answer": {"A"}, "are": {"R"},
    "army": {"TA"}, "at": {"A"}, "bachelor": {"B", "BA"},
    "back": {"B"}, "bad": {"B"}, "bill": {"B"},
    "black": {"B"}, "born": {"B"}, "boy": {"B"},
    "bridge": {"BR"}, "british": {"B", "BR"},
    "carbon": {"C"}, "caught": {"C", "CT"}, "celsius": {"C"},
    "cent": {"C", "P"}, "century": {"C"}, "chapter": {"C", "CH"},
    "church": {"CE", "CH"}, "clubs": {"C"}, "cold": {"C"},
    "communist": {"RED"}, "conservative": {"C"},
    "copper": {"CU"}, "current": {"AC", "I"},
    "daughter": {"D"}, "day": {"D"}, "dead": {"D"},
    "degree": {"D"}, "democrat": {"D"}, "diamonds": {"D"},
    "direction": {"N", "S", "E", "W"}, "doctor": {"DR", "MB", "GP"},
    "duke": {"D"}, "east": {"E"}, "eastern": {"E"},
    "energy": {"E"}, "engineer": {"RE"},
    "english": {"E"}, "european": {"E"},
    "example": {"EG"}, "exercise": {"PE", "PT"},
    "father": {"FR"}, "female": {"F"}, "fifty": {"L"},
    "fine": {"F"}, "first": {"I", "IST"}, "five": {"V"},
    "following": {"F"}, "foot": {"FT"}, "for": {"F"},
    "force": {"F", "G"}, "four": {"IV"}, "french": {"F"},
    "gold": {"AU", "OR"}, "good": {"G"},
    "graduate": {"BA", "MA"}, "grand": {"G", "K"},
    "had": {"D"}, "hard": {"H"}, "has": {"S"},
    "have": {"V"}, "he": {"HE"}, "head": {"H"},
    "hearts": {"H"}, "henry": {"H", "HAL"},
    "hospital": {"H"}, "hot": {"H"}, "hotel": {"H"},
    "hour": {"H", "HR"}, "hundred": {"C"},
    "husband": {"H"}, "hydrogen": {"H"},
    "in": {"IN"}, "international": {"I", "INT"},
    "iron": {"FE"}, "island": {"I"},
    "judge": {"J"}, "junction": {"J", "T"},
    "key": {"K"}, "king": {"K", "R"},
    "knight": {"N", "K"}, "lake": {"L"},
    "large": {"L"}, "last": {"Z"}, "latin": {"L"},
    "lead": {"PB"}, "learner": {"L"}, "left": {"L"},
    "liberal": {"L"}, "line": {"L"}, "litre": {"L"},
    "love": {"O"}, "male": {"M"}, "many": {"C", "D", "M"},
    "mark": {"M"}, "married": {"M"}, "mass": {"M"},
    "master": {"M"}, "member": {"MP"}, "men": {"OR"},
    "midnight": {"M"}, "model": {"T"},
    "money": {"M"}, "morning": {"AM"},
    "motorway": {"M", "MI"}, "name": {"N"},
    "navy": {"RN"}, "new": {"N"}, "nil": {"O"},
    "nitrogen": {"N"}, "noon": {"N"},
    "north": {"N"}, "northern": {"N"},
    "not": {"NT"}, "note": {"N", "DO", "RE", "MI", "FA", "SO", "LA", "TI", "TE"},
    "nothing": {"O", "NIL"}, "number": {"N", "NO"},
    "old": {"O", "EX"}, "one": {"I", "A", "AN"},
    "opening": {"O"}, "oxygen": {"O"},
    "page": {"P"}, "park": {"P"}, "parking": {"P"},
    "party": {"DO"}, "penny": {"P", "D"},
    "phosphorus": {"P"}, "piano": {"P"},
    "point": {"N", "S", "E", "W", "PT"},
    "pole": {"N", "S"}, "political": {"P"},
    "port": {"L", "P"}, "power": {"P"},
    "pressure": {"P"}, "prince": {"P"},
    "princess": {"DI"}, "pupil": {"L"},
    "quarter": {"N", "S", "E", "W"}, "queen": {"R", "ER", "Q"},
    "question": {"Q"}, "quiet": {"P", "SH"},
    "reckoning": {"RE"}, "record": {"EP", "LP"},
    "resistance": {"R"}, "right": {"R", "RT"},
    "ring": {"O"}, "river": {"R"},
    "road": {"RD", "ST"}, "round": {"O"},
    "royal": {"R"}, "run": {"R"},
    "runs": {"R"}, "sailor": {"AB", "TAR"},
    "saint": {"S", "ST"}, "second": {"S", "MO"},
    "silver": {"AG"}, "singular": {"S"},
    "small": {"S"}, "society": {"S"},
    "soldier": {"GI", "RE"}, "son": {"S"},
    "south": {"S"}, "southern": {"S"},
    "spades": {"S"}, "special": {"S"},
    "square": {"S", "T"}, "start": {"S"},
    "stone": {"ST"}, "street": {"ST"},
    "student": {"L"}, "succeeded": {"S"},
    "success": {"S"}, "sulphur": {"S"},
    "sun": {"S"}, "ten": {"X"},
    "the": {"T"}, "thousand": {"K", "M"},
    "time": {"T"}, "tin": {"SN"},
    "top": {"T"}, "trump": {"D"},
    "try": {"GO"}, "tungsten": {"W"},
    "two": {"II"}, "uniform": {"U"},
    "united": {"U"}, "universal": {"U"},
    "university": {"U"}, "unknown": {"X", "Y"},
    "uranium": {"U"}, "very": {"V"},
    "victoria": {"V"}, "victory": {"V"},
    "volt": {"V"}, "volume": {"V"},
    "vote": {"X"}, "was": {"S"},
    "week": {"W"}, "weight": {"W"},
    "well": {"W"}, "west": {"W"},
    "western": {"W"}, "whiskey": {"W"},
    "wicket": {"W"}, "wide": {"W"},
    "wife": {"W"}, "with": {"W"},
    "women": {"W"}, "workers": {"TU"},
    "yard": {"Y"}, "year": {"Y"},
    "zero": {"O"}, "zinc": {"ZN"},
}


def load_pairs():
    with open(INPUT_PATH) as f:
        return [json.loads(line) for line in f]


def is_known_abbrev(word, letters):
    """Check if a single/double letter mapping is a known cryptic abbreviation."""
    w = word.lower().strip()
    l = letters.upper().strip()
    if w in KNOWN_ABBREVS and l in KNOWN_ABBREVS[w]:
        return True
    return False


def filter_pairs(pairs):
    """Apply mechanical filters. Returns (accepted, rejected_with_reasons)."""
    # Load reference data into memory for speed
    print("Loading reference DB into memory...")
    ref = sqlite3.connect(CRYPTIC_DB, timeout=30)
    syn_set = set()
    for row in ref.execute("SELECT LOWER(word), UPPER(synonym) FROM synonyms_pairs"):
        syn_set.add((row[0], row[1]))
    def_set = set()
    for row in ref.execute("SELECT LOWER(definition), UPPER(answer) FROM definition_answers_augmented"):
        def_set.add((row[0], row[1]))
    ref.close()
    print(f"  {len(syn_set)} synonyms, {len(def_set)} definitions loaded")

    clues = sqlite3.connect(CLUES_DB, timeout=30)
    rej_set = set()
    for row in clues.execute("SELECT LOWER(word), UPPER(letters) FROM rejected_enrichments"):
        rej_set.add((row[0], row[1]))
    clues.close()
    print(f"  {len(rej_set)} rejected pairs loaded")

    accepted = []
    rejected = []

    for p in pairs:
        word = p["word"].strip().lower()
        letters = p["letters"].strip().upper()
        reason = None

        # 1. Circular: word == letters
        if word.upper() == letters:
            reason = "circular"

        # 2. Already in synonyms_pairs
        elif (word, letters) in syn_set:
            reason = "already_synonym"

        # 3. Already in definition_answers_augmented
        elif (word, letters) in def_set:
            reason = "already_definition"

        # 4. Previously rejected
        elif (word, letters) in rej_set:
            reason = "previously_rejected"

        # 5. Single letter that isn't a known abbreviation
        elif len(letters) == 1 and not is_known_abbrev(word, letters):
            reason = "unknown_single_letter"

        # 6. Two letters that isn't a known abbreviation and word is very common
        elif len(letters) == 2 and not is_known_abbrev(word, letters):
            if len(word) <= 3:
                reason = "short_word_to_2_letters"

        # 7. Letters contained entirely within the word (extraction artefact)
        elif len(letters) >= 2 and letters.lower() in word:
            reason = "letters_in_word"

        # 8. Word contained entirely within the letters (likely reverse)
        elif len(word) >= 3 and word.upper() in letters:
            reason = "word_in_letters"

        # 9. Hyphenated compound words where one part IS the letters
        elif "-" in word:
            parts = word.split("-")
            if any(part.upper() == letters for part in parts):
                reason = "hyphen_part_is_letters"

        if reason:
            rejected.append((word, letters, reason))
        else:
            accepted.append(p)

    return accepted, rejected


def export_filtered(accepted):
    """Export filtered pairs to Excel for final review."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enrichment Review"

    headers = ["Word", "Letters", "Decision"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Sort by word for easier review
    accepted.sort(key=lambda p: p["word"].lower())

    for i, p in enumerate(accepted, 2):
        ws.cell(row=i, column=1, value=p["word"]).font = Font(size=12)
        ws.cell(row=i, column=2, value=p["letters"]).font = Font(size=12, bold=True)
        ws.cell(row=i, column=3, value="").font = Font(size=12)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    print(f"Exported {len(accepted)} pairs to {OUTPUT_PATH}")


def main():
    pairs = load_pairs()
    print(f"Loaded {len(pairs)} Haiku-validated pairs")

    accepted, rejected = filter_pairs(pairs)

    # Summary
    from collections import Counter
    reasons = Counter(r for _, _, r in rejected)
    print(f"\nRejected: {len(rejected)}")
    for reason, count in reasons.most_common():
        print(f"  {reason}: {count}")
    print(f"\nAccepted for review: {len(accepted)}")

    export_filtered(accepted)
    print(f"\nReview in Excel, mark column C with Y or N")


if __name__ == "__main__":
    main()
