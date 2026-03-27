"""Strict Haiku review of enrichment pairs.

Sends pairs to Haiku with a strict prompt asking:
- Is this a valid synonym, abbreviation, or definition IN CRYPTIC CROSSWORDS?
- Would a cryptic setter use this mapping?

Only keeps pairs Haiku rates as valid.
"""

import json
import os
import re
import sys
import time

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from dotenv import load_dotenv
load_dotenv(os.path.join(ROOT, ".env"))

INPUT_PATH = os.path.join(ROOT, "data", "enrichment_review_filtered.xlsx")
OUTPUT_PATH = os.path.join(ROOT, "data", "enrichment_final_review.xlsx")


def load_pairs():
    import openpyxl
    wb = openpyxl.load_workbook(INPUT_PATH)
    ws = wb.active
    pairs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        word, letters, _ = row
        if word and letters:
            pairs.append({"word": str(word).strip(), "letters": str(letters).strip()})
    return pairs


def verify_batch(client, pairs):
    """Send a batch to Haiku for strict verification."""
    pair_lines = [f"{p['word']} -> {p['letters']}" for p in pairs]

    prompt = f"""You are a strict cryptic crossword expert. For each pair below, determine if the left side
can legitimately lead to the right side in a cryptic crossword clue.

Valid reasons to accept:
- Direct synonym (e.g. "trouble" -> "CONCERN")
- Standard cryptic abbreviation (e.g. "doctor" -> "DR")
- Definition (e.g. "gardens" -> "KEW")
- Cryptic convention (e.g. "flower" -> "RIVER" because a river flows)

REJECT if:
- The connection is too tenuous or requires multiple leaps
- It's a random word association, not a crossword convention
- The word is a number/code and the mapping makes no sense
- It's a partial word or extraction artefact
- You can't explain how a setter would use this mapping

Be STRICT. When in doubt, reject. Better to miss a valid pair than accept a wrong one.

Reply with ONLY a JSON array. Each object: {{"word": "...", "letters": "...", "valid": true/false}}

Pairs:
{chr(10).join(pair_lines)}"""

    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=4096,
        temperature=0,
        messages=[{"role": "user", "content": prompt}],
    )

    try:
        text = response.content[0].text
        json_match = re.search(r"\[.*\]", text, re.DOTALL)
        if json_match:
            return json.loads(json_match.group())
    except Exception as e:
        print(f"  Parse error: {e}")
    return []


def main():
    import anthropic

    pairs = load_pairs()
    print(f"Loaded {len(pairs)} pairs from filtered Excel")

    client = anthropic.Anthropic()
    chunk_size = 50  # Smaller chunks for more careful review
    all_valid = []
    all_invalid = []

    for i in range(0, len(pairs), chunk_size):
        chunk = pairs[i:i + chunk_size]
        print(f"Verifying {i+1}-{i+len(chunk)} of {len(pairs)}...")

        results = verify_batch(client, chunk)
        valid = [r for r in results if r.get("valid")]
        invalid = [r for r in results if not r.get("valid")]
        all_valid.extend(valid)
        all_invalid.extend(invalid)

        n_valid = len(valid)
        n_invalid = len(invalid)
        print(f"  {n_valid} accepted, {n_invalid} rejected")
        time.sleep(0.5)  # Rate limiting

    print(f"\n{'='*60}")
    print(f"STRICT REVIEW: {len(all_valid)}/{len(pairs)} accepted ({100*len(all_valid)//len(pairs)}%)")
    print(f"Rejected: {len(all_invalid)}")
    print(f"{'='*60}")

    # Export to Excel
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Final Review"

    headers = ["Word", "Letters", "Decision"]
    header_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    all_valid.sort(key=lambda p: p["word"].lower())
    for i, p in enumerate(all_valid, 2):
        ws.cell(row=i, column=1, value=p["word"]).font = Font(size=12)
        ws.cell(row=i, column=2, value=p["letters"]).font = Font(size=12, bold=True)
        ws.cell(row=i, column=3, value="").font = Font(size=12)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    print(f"\nExported {len(all_valid)} pairs to {OUTPUT_PATH}")
    print("Review in Excel, mark column C with Y (accept) or N (reject)")

    # Also save rejected for reference
    rejected_path = os.path.join(ROOT, "data", "enrichment_haiku_rejected.jsonl")
    with open(rejected_path, "w") as f:
        for r in all_invalid:
            f.write(json.dumps(r) + "\n")
    print(f"Rejected pairs saved to {rejected_path}")


if __name__ == "__main__":
    main()
